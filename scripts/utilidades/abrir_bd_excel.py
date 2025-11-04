import os
import sys
from openpyxl import load_workbook
from datetime import datetime

lista_archivos = [
    r"C:\Users\ELOPEZ21334\OneDrive - Distribuidora La Florida S.A\Proyectos Reportes 3PL\INCIDENCIAS\INCIDENCIAS_2025.xlsx",
    r"C:\Users\ELOPEZ21334\OneDrive - Distribuidora La Florida S.A\Proyectos Reportes 3PL\BO HISTORICO\BO - HISTORICO PARA POWER BI 2025.xlsx",
    r"C:\Users\ELOPEZ21334\OneDrive - Distribuidora La Florida S.A\Proyectos Reportes 3PL\3-Reporte de Tipificación de Devoluciones\País\2025\10-2025-Devoluciones.xlsx",
    r"C:\Users\ELOPEZ21334\OneDrive - Distribuidora La Florida S.A\Retail\Proyectos de Reportes\2023\OTIF ENT CD01\REP PLR ESTATUS ENTREGAS v25.xlsx",
    r"C:\Users\ELOPEZ21334\OneDrive - Distribuidora La Florida S.A\Retail\Proyectos de Reportes\2023\Torre de Control\YTD 2025\VOL POR PORTAFOLIO ENE-2025",  # <-- SIN EXTENSIÓN
    r"C:\Users\ELOPEZ21334\OneDrive - Distribuidora La Florida S.A\Temporada 2024\MONITOR GUIA.xlsx"
]

log_file = r"C:\Users\ELOPEZ21334\OneDrive - Distribuidora La Florida S.A\Proyectos Reportes 3PL\log.txt"

def info_archivo(path):
    try:
        size = os.path.getsize(path)
        mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
        return f" [size={size} bytes, mtime={mtime}]"
    except Exception:
        return ""

def log_write(log, mensaje):
    log.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {mensaje}\n")

with open(log_file, "a", encoding="utf-8") as log:
    log_write(log, "===== INICIO DE VALIDACIÓN =====")
    for archivo in lista_archivos:
        try:
            print(f"Validando: {archivo}")
            # 1) ¿Existe?
            if not os.path.exists(archivo):
                log_write(log, f"NO ENCONTRADO: {archivo}")
                continue

            # 2) ¿Es carpeta?
            if os.path.isdir(archivo):
                log_write(log, f"ES UNA CARPETA (no un archivo): {archivo}")
                continue

            # 3) Validar extensión
            ext = os.path.splitext(archivo)[1].lower()
            if ext not in [".xlsx", ".xlsm"]:
                log_write(log, f"EXTENSIÓN NO SOPORTADA ({ext}): {archivo}{info_archivo(archivo)}")
                continue

            # 4) Intentar abrir
            try:
                wb = load_workbook(archivo, read_only=True, data_only=True)
                # Opcional: leer nombres de hojas para confirmar
                sheets = ", ".join(wb.sheetnames[:5])
                wb.close()
                log_write(log, f"OK: {archivo}{info_archivo(archivo)} | Hojas: {sheets}")
            except PermissionError as e:
                log_write(log, f"BLOQUEADO / PERMISO DENEGADO: {archivo}{info_archivo(archivo)} | Detalle: {e}")
            except Exception as e:
                # Mensaje típico si está corrupto o protegido
                log_write(log, f"ERROR AL ABRIR: {archivo}{info_archivo(archivo)} | Detalle: {repr(e)}")

        except Exception as e:
            log_write(log, f"ERROR NO CONTROLADO con {archivo}: {repr(e)}")

    log_write(log, "===== FIN DE VALIDACIÓN =====")