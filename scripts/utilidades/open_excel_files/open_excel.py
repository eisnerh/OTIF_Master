import os
import sys
import time
import json
import ctypes
import subprocess
from datetime import datetime
from openpyxl import load_workbook

# ================== CONFIGURACION ==================
CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")

def load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

cfg = load_config(CONFIG_PATH)
ARCHIVOS = cfg.get("archivos", [])
LOG_FILE = cfg.get("log_file")

# Configuraciones
MAX_DOWNLOAD_WAIT = 60  # segundos para esperar descarga
CHECK_INTERVAL = 2  # segundos entre verificaciones

# ================== ATRIBUTOS WINDOWS (OneDrive) ==================
FILE_ATTRIBUTE_OFFLINE = 0x1000
FILE_ATTRIBUTE_RECALL_ON_OPEN = 0x40000
FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS = 0x400000

def get_file_attrs(path):
    """Obtiene los atributos del archivo en Windows"""
    if os.name != "nt":
        return 0
    try:
        GetFileAttributesW = ctypes.windll.kernel32.GetFileAttributesW
        GetFileAttributesW.argtypes = [ctypes.c_wchar_p]
        GetFileAttributesW.restype = ctypes.c_uint32
        attrs = GetFileAttributesW(path)
        if attrs == 0xFFFFFFFF:
            return 0
        return attrs
    except Exception:
        return 0

def is_onedrive_online_only(path):
    """Detecta si el archivo esta en modo online-only de OneDrive"""
    attrs = get_file_attrs(path)
    return bool(attrs & (FILE_ATTRIBUTE_RECALL_ON_OPEN | FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS))

def trigger_download(path):
    """Activa la descarga de un archivo de OneDrive y espera a que termine"""
    print(f"  >> Archivo no descargado. Iniciando descarga desde OneDrive...")
    
    # Intentar leer el archivo para activar la descarga
    try:
        with open(path, 'rb') as f:
            f.read(1)  # Leer solo 1 byte para activar descarga
    except Exception:
        pass
    
    # Esperar a que se descargue
    start_time = time.time()
    while time.time() - start_time < MAX_DOWNLOAD_WAIT:
        if not is_onedrive_online_only(path):
            size = os.path.getsize(path)
            print(f"  >> Descarga completada ({size:,} bytes)")
            return True
        time.sleep(CHECK_INTERVAL)
        print(f"  >> Esperando descarga... ({int(time.time() - start_time)}s)")
    
    print(f"  >> ADVERTENCIA: Tiempo de espera agotado")
    return False

def verify_can_open(path):
    """Verifica si el archivo se puede abrir con openpyxl"""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return True, sheets
    except Exception as e:
        return False, str(e)

def open_with_excel(path):
    """Abre el archivo con Excel (programa predeterminado)"""
    try:
        if os.name == 'nt':
            os.startfile(path)
        elif sys.platform == 'darwin':
            subprocess.call(['open', path])
        else:
            subprocess.call(['xdg-open', path])
        return True
    except Exception as e:
        return False

def log_write(mensaje):
    """Escribe en consola y archivo de log"""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{ts} - {mensaje}"
    print(line)
    if LOG_FILE:
        try:
            with open(LOG_FILE, "a", encoding="utf-8") as log:
                log.write(line + "\n")
        except Exception:
            pass

def main():
    log_write("=" * 60)
    log_write("INICIO - Abriendo archivos Excel")
    log_write("=" * 60)
    
    archivos_abiertos = 0
    archivos_error = 0
    
    for idx, archivo in enumerate(ARCHIVOS, start=1):
        print()
        log_write(f"[{idx}/{len(ARCHIVOS)}] Procesando: {os.path.basename(archivo)}")
        
        # Verificar existencia
        if not os.path.exists(archivo):
            log_write(f"  >> ERROR: Archivo no encontrado")
            archivos_error += 1
            continue
        
        # Verificar si esta online-only y descargar si es necesario
        if is_onedrive_online_only(archivo):
            log_write(f"  >> Archivo en modo online-only de OneDrive")
            if not trigger_download(archivo):
                log_write(f"  >> ERROR: No se pudo descargar completamente")
                archivos_error += 1
                continue
        else:
            size = os.path.getsize(archivo)
            log_write(f"  >> Archivo disponible localmente ({size:,} bytes)")
        
        # Verificar que se puede abrir
        log_write(f"  >> Verificando integridad del archivo...")
        can_open, result = verify_can_open(archivo)
        
        if not can_open:
            log_write(f"  >> ERROR: No se puede abrir - {result}")
            archivos_error += 1
            continue
        
        log_write(f"  >> Archivo valido - Hojas: {', '.join(result[:3])}{'...' if len(result) > 3 else ''}")
        
        # Abrir con Excel
        log_write(f"  >> Abriendo con Excel...")
        if open_with_excel(archivo):
            log_write(f"  >> OK: Archivo abierto exitosamente")
            archivos_abiertos += 1
        else:
            log_write(f"  >> ERROR: No se pudo abrir con Excel")
            archivos_error += 1
        
        # Pausa entre archivos (excepto el ultimo)
        if idx < len(ARCHIVOS):
            time.sleep(1)
    
    print()
    log_write("=" * 60)
    log_write(f"RESUMEN: {archivos_abiertos} archivos abiertos, {archivos_error} errores")
    log_write("=" * 60)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nProceso interrumpido por el usuario")
    except Exception as e:
        log_write(f"ERROR FATAL: {repr(e)}")
        print(f"\nERROR FATAL: {repr(e)}")

