import os
import sys
import time
import json
import csv
import zipfile
import ctypes
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile

# ================== LECTURA CONFIG ==================
CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")

def load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

cfg = load_config(CONFIG_PATH)

ARCHIVOS = cfg.get("archivos", [])
LOG_FILE = cfg.get("log_file")
CSV_RESUMEN = cfg.get("csv_resumen")

SLEEP_BETWEEN_FILES = float(cfg.get("sleep_between_files", 2.0))
RETRY_OPEN_ATTEMPTS = int(cfg.get("retry_open_attempts", 4))
RETRY_BASE_DELAY = float(cfg.get("retry_base_delay", 1.5))
CHECK_LOCK_TIMEOUT = float(cfg.get("check_lock_timeout", 8.0))
CHECK_LOCK_INTERVAL = float(cfg.get("check_lock_interval", 0.5))

# ================== LOGGING ==================
def log_write(log, mensaje):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{ts} - {mensaje}\n"
    log.write(line)
    log.flush()
    print(line, end="")

def info_archivo(path):
    try:
        size = os.path.getsize(path)
        mtime = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
        return f" [size={size} bytes, mtime={mtime}]"
    except Exception:
        return ""

# ================== ATRIBUTOS WINDOWS (OneDrive/Cloud) ==================
# https://learn.microsoft.com/en-us/windows/win32/fileio/file-attribute-constants
FILE_ATTRIBUTE_READONLY = 0x0001
FILE_ATTRIBUTE_HIDDEN = 0x0002
FILE_ATTRIBUTE_SYSTEM = 0x0004
FILE_ATTRIBUTE_DIRECTORY = 0x0010
FILE_ATTRIBUTE_ARCHIVE = 0x0020
FILE_ATTRIBUTE_NORMAL = 0x0080
FILE_ATTRIBUTE_TEMPORARY = 0x0100
FILE_ATTRIBUTE_SPARSE_FILE = 0x0200
FILE_ATTRIBUTE_REPARSE_POINT = 0x0400
FILE_ATTRIBUTE_COMPRESSED = 0x0800
FILE_ATTRIBUTE_OFFLINE = 0x1000
FILE_ATTRIBUTE_NOT_CONTENT_INDEXED = 0x2000
FILE_ATTRIBUTE_ENCRYPTED = 0x4000
FILE_ATTRIBUTE_INTEGRITY_STREAM = 0x8000
FILE_ATTRIBUTE_VIRTUAL = 0x10000
FILE_ATTRIBUTE_NO_SCRUB_DATA = 0x20000
FILE_ATTRIBUTE_PINNED = 0x80000
FILE_ATTRIBUTE_UNPINNED = 0x100000
FILE_ATTRIBUTE_RECALL_ON_OPEN = 0x40000  # alias compartido con EA en algunas docs
FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS = 0x400000

def get_file_attrs(path):
    """
    Devuelve (attrs_int, attrs_list_legibles) en Windows; vacío en otros SO.
    """
    if os.name != "nt":
        return (0, [])
    GetFileAttributesW = ctypes.windll.kernel32.GetFileAttributesW
    GetFileAttributesW.argtypes = [ctypes.c_wchar_p]
    GetFileAttributesW.restype = ctypes.c_uint32
    attrs = GetFileAttributesW(path)
    if attrs == 0xFFFFFFFF:
        return (0, [])
    flags = []
    mapping = [
        (FILE_ATTRIBUTE_READONLY, "READONLY"),
        (FILE_ATTRIBUTE_HIDDEN, "HIDDEN"),
        (FILE_ATTRIBUTE_SYSTEM, "SYSTEM"),
        (FILE_ATTRIBUTE_DIRECTORY, "DIRECTORY"),
        (FILE_ATTRIBUTE_ARCHIVE, "ARCHIVE"),
        (FILE_ATTRIBUTE_NORMAL, "NORMAL"),
        (FILE_ATTRIBUTE_TEMPORARY, "TEMPORARY"),
        (FILE_ATTRIBUTE_SPARSE_FILE, "SPARSE"),
        (FILE_ATTRIBUTE_REPARSE_POINT, "REPARSE_POINT"),
        (FILE_ATTRIBUTE_COMPRESSED, "COMPRESSED"),
        (FILE_ATTRIBUTE_OFFLINE, "OFFLINE"),
        (FILE_ATTRIBUTE_NOT_CONTENT_INDEXED, "NOT_INDEXED"),
        (FILE_ATTRIBUTE_ENCRYPTED, "ENCRYPTED"),
        (FILE_ATTRIBUTE_INTEGRITY_STREAM, "INTEGRITY"),
        (FILE_ATTRIBUTE_VIRTUAL, "VIRTUAL"),
        (FILE_ATTRIBUTE_NO_SCRUB_DATA, "NO_SCRUB"),
        (FILE_ATTRIBUTE_PINNED, "PINNED"),
        (FILE_ATTRIBUTE_UNPINNED, "UNPINNED"),
        (FILE_ATTRIBUTE_RECALL_ON_OPEN, "RECALL_ON_OPEN"),
        (FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS, "RECALL_ON_DATA_ACCESS"),
    ]
    for bit, name in mapping:
        if attrs & bit:
            flags.append(name)
    return (attrs, flags)

def is_file_locked(path, timeout=CHECK_LOCK_TIMEOUT, interval=CHECK_LOCK_INTERVAL):
    """
    Detecta bloqueo compartido intentando abrir el descriptor.
    """
    if not os.path.exists(path) or os.path.isdir(path):
        return False
    start = time.time()
    flags = os.O_RDONLY
    if os.name == "nt":
        flags |= os.O_BINARY
    while time.time() - start < timeout:
        try:
            fd = os.open(path, flags)
            os.close(fd)
            return False
        except PermissionError:
            time.sleep(interval)
        except OSError:
            time.sleep(interval)
    return True

def looks_like_com_ole_timeout(exc: Exception) -> bool:
    msg = str(exc).upper()
    needles = [
        "RPC_E_SERVERCALL_RETRYLATER",
        "RPC_E_CALL_REJECTED",
        "CALL WAS REJECTED BY CALLEE",
        "OLE",
        "COM_ERROR",
        "AUTOMATION ERROR"
    ]
    return any(n in msg for n in needles)

def probe_zip_encryption_and_integrity(path):
    """
    Devuelve (status, detail)
    status:
      - 'zip_ok'
      - 'encrypted' (xlsx cifrado con contraseña)
      - 'bad_zip' (corrupción)
      - 'not_zip' (no parece zip/xlsx válido)
    """
    try:
        with zipfile.ZipFile(path, 'r') as zf:
            names = zf.namelist()
            # XLSX cifrado típico: contiene 'EncryptionInfo' y 'EncryptedPackage'
            if any("EncryptionInfo" in n for n in names) and any("EncryptedPackage" in n for n in names):
                return ("encrypted", "Se detectó estructura de cifrado (EncryptionInfo/EncryptedPackage)")
            if not any(n.startswith("xl/") for n in names):
                return ("not_zip", "El contenedor ZIP no contiene carpeta xl/ (formato inesperado)")
            return ("zip_ok", f"{len(names)} entradas ZIP")
    except BadZipFile as e:
        return ("bad_zip", f"BadZipFile: {e}")
    except Exception as e:
        return ("not_zip", f"Error leyendo ZIP: {repr(e)}")

def open_workbook_with_retries(path, log, attempts=RETRY_OPEN_ATTEMPTS, base_delay=RETRY_BASE_DELAY):
    last_exc = None
    for retry in range(attempts):
        if is_file_locked(path):
            log_write(log, f"ARCHIVO BLOQUEADO (esperando): {path}{info_archivo(path)}")
        else:
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                return wb
            except PermissionError as e:
                last_exc = e
                log_write(log, f"PERMISO/BLOQUEO al abrir (intento {retry+1}/{attempts}): {path} | Detalle: {e}")
            except InvalidFileException as e:
                last_exc = e
                raise
            except BadZipFile as e:
                last_exc = e
                raise
            except Exception as e:
                last_exc = e
                if looks_like_com_ole_timeout(e):
                    log_write(log, f"POSIBLE OLE/COM TIMEOUT al abrir (intento {retry+1}/{attempts}): {path} | Detalle: {repr(e)}")
                else:
                    log_write(log, f"ERROR AL ABRIR (intento {retry+1}/{attempts}): {path} | Detalle: {repr(e)}")
        time.sleep(base_delay * (2 ** retry))
    if last_exc:
        raise last_exc
    raise RuntimeError("Fallo desconocido al abrir el libro.")

def extension_valida(path):
    ext = os.path.splitext(path)[1].lower()
    return ext in (".xlsx", ".xlsm")

def write_csv_header(csv_path):
    new = not os.path.exists(csv_path)
    f = open(csv_path, "a", newline="", encoding="utf-8")
    writer = csv.writer(f)
    if new:
        writer.writerow([
            "archivo", "estado", "detalle", "ext", "size_bytes", "mtime",
            "sheets_preview", "attrs", "zip_probe"
        ])
    return f, writer

def main():
    ok_count = 0
    with open(LOG_FILE, "a", encoding="utf-8") as log:
        log_write(log, "===== INICIO DE VALIDACIÓN =====")
        log_write(log, f"Python: {sys.version}")
        try:
            import openpyxl
            log_write(log, f"openpyxl: {openpyxl.__version__}")
        except Exception:
            pass

        csv_f, csv_w = write_csv_header(CSV_RESUMEN)

        for idx, archivo in enumerate(ARCHIVOS, start=1):
            try:
                if idx > 1 and SLEEP_BETWEEN_FILES > 0:
                    time.sleep(SLEEP_BETWEEN_FILES)

                log_write(log, f"Validando: {archivo}")

                if not os.path.exists(archivo):
                    log_write(log, f"NO ENCONTRADO: {archivo}")
                    csv_w.writerow([archivo, "NO_ENCONTRADO", "", "", "", "", "", "", ""])
                    continue

                if os.path.isdir(archivo):
                    log_write(log, f"ES UNA CARPETA (no archivo): {archivo}")
                    csv_w.writerow([archivo, "ES_CARPETA", "", "", "", "", "", "", ""])
                    continue

                ext = os.path.splitext(archivo)[1].lower()
                size = os.path.getsize(archivo)
                mtime = datetime.fromtimestamp(os.path.getmtime(archivo)).strftime("%Y-%m-%d %H:%M:%S")
                attrs_int, attrs_list = get_file_attrs(archivo)
                attrs_str = ",".join(attrs_list) if attrs_list else ""

                if not extension_valida(archivo):
                    log_write(log, f"EXTENSIÓN NO SOPORTADA ({ext}): {archivo}{info_archivo(archivo)}")
                    csv_w.writerow([archivo, "EXT_NO_SOPORTADA", f"{ext}", ext, size, mtime, "", attrs_str, ""])
                    continue

                # Sonda ZIP para cifrado/corrupción
                zip_status, zip_detail = probe_zip_encryption_and_integrity(archivo)
                if zip_status == "encrypted":
                    log_write(log, f"ARCHIVO CIFRADO (password): {archivo}{info_archivo(archivo)} | {zip_detail}")
                    csv_w.writerow([archivo, "CIFRADO", zip_detail, ext, size, mtime, "", attrs_str, zip_status])
                    continue
                elif zip_status == "bad_zip":
                    log_write(log, f"ARCHIVO CORRUPTO (ZIP): {archivo}{info_archivo(archivo)} | {zip_detail}")
                    csv_w.writerow([archivo, "CORRUPTO_ZIP", zip_detail, ext, size, mtime, "", attrs_str, zip_status])
                    continue
                elif zip_status == "not_zip":
                    log_write(log, f"NO ES ZIP/FORMATO XLSX VÁLIDO: {archivo}{info_archivo(archivo)} | {zip_detail}")
                    csv_w.writerow([archivo, "FORMATO_INVALIDO", zip_detail, ext, size, mtime, "", attrs_str, zip_status])
                    continue

                try:
                    wb = open_workbook_with_retries(archivo, log, attempts=RETRY_OPEN_ATTEMPTS, base_delay=RETRY_BASE_DELAY)
                    sheets = ", ".join(wb.sheetnames[:5])
                    wb.close()
                    log_write(log, f"OK: {archivo}{info_archivo(archivo)} | Hojas: {sheets} | Attrs: {attrs_str}")
                    csv_w.writerow([archivo, "OK", "", ext, size, mtime, sheets, attrs_str, zip_status])
                    ok_count += 1
                except PermissionError as e:
                    log_write(log, f"BLOQUEADO / PERMISO DENEGADO (tras reintentos): {archivo}{info_archivo(archivo)} | {e}")
                    csv_w.writerow([archivo, "BLOQUEADO", str(e), ext, size, mtime, "", attrs_str, zip_status])
                except InvalidFileException as e:
                    log_write(log, f"ARCHIVO NO VÁLIDO/NO SOPORTADO: {archivo}{info_archivo(archivo)} | {repr(e)}")
                    csv_w.writerow([archivo, "INVALIDO_OPENPYXL", repr(e), ext, size, mtime, "", attrs_str, zip_status])
                except BadZipFile as e:
                    log_write(log, f"ARCHIVO CORRUPTO (ZIP) al abrir: {archivo}{info_archivo(archivo)} | {repr(e)}")
                    csv_w.writerow([archivo, "CORRUPTO_ZIP", repr(e), ext, size, mtime, "", attrs_str, zip_status])
                except Exception as e:
                    if looks_like_com_ole_timeout(e):
                        log_write(log, f"POSIBLE OLE/COM TIMEOUT (tras reintentos): {archivo}{info_archivo(archivo)} | {repr(e)}")
                        csv_w.writerow([archivo, "POSIBLE_OLE_TIMEOUT", repr(e), ext, size, mtime, "", attrs_str, zip_status])
                    else:
                        log_write(log, f"ERROR AL ABRIR: {archivo}{info_archivo(archivo)} | {repr(e)}")
                        csv_w.writerow([archivo, "ERROR", repr(e), ext, size, mtime, "", attrs_str, zip_status])

            except Exception as e:
                log_write(log, f"ERROR NO CONTROLADO con {archivo}: {repr(e)}")
                csv_w.writerow([archivo, "ERROR_NO_CONTROLADO", repr(e), "", "", "", "", "", ""])

        csv_f.close()

        # Autodiagnóstico: si no se abrió ninguno, verificar openpyxl con un archivo local temporal
        if ok_count == 0:
            tmp_dir = os.path.join(os.getenv("TEMP", os.getcwd()), "xlsx_diag")
            os.makedirs(tmp_dir, exist_ok=True)
            test_path = os.path.join(tmp_dir, "test_openpyxl.xlsx")
            log_write(log, f"NINGÚN ARCHIVO ABIERTO. Realizando prueba local: {test_path}")

            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws["A1"] = "Prueba openpyxl"
                wb.save(test_path)
                wb.close()
                wb2 = load_workbook(test_path, read_only=True, data_only=True)
                sheets = ", ".join(wb2.sheetnames)
                wb2.close()
                log_write(log, f"PRUEBA OK: openpyxl funciona localmente | Hojas: {sheets}")
                log_write(log, "=> Conclusión: la falla apunta a accesibilidad de archivos (bloqueo, OneDrive online-only, cifrado, permisos).")
            except Exception as e:
                log_write(log, f"PRUEBA FALLÓ: openpyxl no pudo operar localmente | {repr(e)}")
                log_write(log, "=> Conclusión: revisar instalación/entorno de Python y openpyxl.")

        log_write(log, "===== FIN DE VALIDACIÓN =====")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # Fallback de error fatal
        try:
            with open(LOG_FILE, "a", encoding="utf-8") as log:
                log_write(log, f"ERROR FATAL: {repr(e)}")
        except Exception:
            print(f"ERROR FATAL: {repr(e)}")